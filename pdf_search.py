import os
import re
import fitz  # PyMuPDF
import pandas as pd
import argparse
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse
import json

# === REGEX ===
url_regex = re.compile(r"https?://[^\s<>)\"']+")
# Generic Bates-style reference number: a short letter prefix followed by a
# padded run of digits (e.g. "ACME-000123", "SMITH_0001234", "ABC000123"),
# optionally followed by a dotted page (or page range) suffix, such as
#   ACME0000123.8              -> single page offset
#   ACME0000123.25-26          -> range of page offsets
#   ACME0000456.542-550        -> range of page offsets
# Group 1 = base Bates ID; group 2 = optional ".suffix" (including the dot);
# group 0 = the full citation as it appeared in the source text.
generic_bates_regex = re.compile(
    r"\b([A-Za-z]{1,10}[_\-]?\d{3,10}\b)"
    r"(\.\d+(?:[-\u2013]\d+)?)"
    r"?"
)

def parse_page_range(suffix):
    """Parse a Bates suffix such as '8' or '25-26' (or '.25-26') into a
    (start, end) tuple of ints. Returns None if the suffix is missing or
    unparseable. Accepts both ASCII hyphens and en/em dashes."""
    if not suffix:
        return None
    s = str(suffix).strip().lstrip(".").strip()
    s = s.replace("\u2013", "-").replace("\u2014", "-")
    if "-" in s:
        start_s, end_s = s.split("-", 1)
    else:
        start_s = end_s = s
    try:
        start = int(start_s)
        end = int(end_s)
    except ValueError:
        return None
    if end < start:
        start, end = end, start
    return (start, end)

def merge_page_ranges(ranges):
    """Merge a list of (start, end) tuples into a minimal, sorted set of
    non-overlapping ranges. Ranges that touch (e.g. (1,5)+(6,10)) are
    combined into a single range (e.g. (1,10))."""
    if not ranges:
        return []
    ordered = sorted((int(s), int(e)) for (s, e) in ranges)
    merged = [list(ordered[0])]
    for s, e in ordered[1:]:
        cur = merged[-1]
        if s <= cur[1] + 1:
            cur[1] = max(cur[1], e)
        else:
            merged.append([s, e])
    return [tuple(r) for r in merged]

def format_merged_ranges(merged):
    """Render a list of (start, end) tuples as a compact string, e.g.
    [(8, 8), (25, 26)] -> '8, 25-26'."""
    parts = []
    for s, e in merged:
        parts.append(str(s) if s == e else f"{s}-{e}")
    return ", ".join(parts)

# Known source code file extensions (used with --source-code)
SOURCE_CODE_EXTENSIONS = [
    "py", "js", "ts", "jsx", "tsx", "c", "h", "cpp", "hpp", "cc", "cs",
    "java", "rb", "go", "rs", "swift", "kt", "kts", "scala", "php",
    "sh", "bash", "ps1", "m", "mm", "lua", "pl", "r",
]

def build_file_pattern(extensions):
    ext_alts = "|".join(re.escape(e) for e in extensions)
    return re.compile(rf"\b[\w.\-/\\]+\.(?:{ext_alts})\b", re.IGNORECASE)

def get_base_url(url):
    try:
        parsed = urlparse(url)
        return parsed.netloc.lower()
    except:
        return ""
    
import os

def get_unique_output_folder(base_name="run"):
    os.makedirs("output", exist_ok=True)
    i = 1
    folder = os.path.join("output", base_name)
    while os.path.exists(folder):
        folder = os.path.join("output", f"{base_name}_{i}")
        i += 1
    os.makedirs(folder)
    return folder

def save_large_json(data, base_filename, folder, threshold=5000):
    """Save large JSON data to external file if size exceeds threshold."""
    json_str = json.dumps(data, indent=2, ensure_ascii=False)

    if len(json_str) <= threshold:
        return json_str

    os.makedirs(folder, exist_ok=True)
    safe_name = re.sub(r'[^\w\-_.]', '_', base_filename)[:80]  # sanitize + truncate
    path = os.path.join(folder, f"{safe_name}.json")

    with open(path, "w", encoding="utf-8") as f:
        f.write(json_str)

    return f"[See {os.path.basename(path)}]"

def parse_args():
    parser = argparse.ArgumentParser(description="Extract links, keywords, and Bates numbers from PDFs.")
    parser.add_argument("path", help="Path to a single PDF file to process (default). Use --folder to instead treat this as a folder of PDFs to scan recursively.")
    parser.add_argument("--folder", action="store_true", help="Treat 'path' as a folder and recursively scan it for PDF files, instead of a single PDF file.")
    parser.add_argument("--output", default="pdf_extraction_output.xlsx", help="Output Excel file path")
    parser.add_argument("--link-annotations", action="store_true", help="Extract embedded link annotations from PDFs")
    parser.add_argument("--text-urls", action="store_true", help="Extract URLs from visible page text")
    parser.add_argument("--keywords", nargs="*", default=[], help="Search for specific keywords")
    parser.add_argument("--keywords-file", help="Path to file with one keyword per line")
    parser.add_argument("--bates-footer-prefix", help="Prefix string to identify Bates numbers in the bottom-right footer of each page (e.g. 'MyCompany')")
    parser.add_argument("--bates-body", action="store_true", help="Search each page's visible text/content for any Bates-style reference numbers (any prefix), not just the footer. Useful for capturing Bates numbers cited from other document sets.")
    parser.add_argument("--source-code", action="store_true", help="Search each page's visible text for citations of common source code file names (e.g. .py, .js, .java, .c, .ts, ...).")
    parser.add_argument("--file-ext", nargs="+", default=[], help="Search each page's visible text for citations of file names with a specific extension (e.g. --file-ext py txt).")
    parser.add_argument("--context-window", type=int, default=100, help="Number of characters prepending and appending matched string that are returned in 'Context' parameter of 'References' column")

    return parser.parse_args()


def extract_context(text, match_start, match_end, window):
    return text[max(0, match_start - window):match_end + window]

def clean_url_string(url: str) -> str:
    # Remove newlines and collapse spaces
    return re.sub(r"\s+", "", url.replace('\n', '').replace('\r', ''))

def clean_context_string(text: str) -> str:
    text = text.replace('\n', ' ').replace('\r', '')
    text = re.sub(r"[•■·]", "", text)
    return re.sub(r'\s+', ' ', text).strip()

def format_excel(filepath):
    wb = load_workbook(filepath)
    for ws in wb.worksheets:
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Wrap text and auto-fit width
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True)
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    wb.save(filepath)
    
def format_references_to_json(refs):
    return json.dumps(refs, indent=2, ensure_ascii=False)

def main():
    args = parse_args()
    output_dir = get_unique_output_folder(os.path.splitext(os.path.basename(args.output))[0])
    excel_path = os.path.join(output_dir, os.path.basename(args.output))
    
    # Load keywords from file if specified
    if args.keywords_file:
        try:
            with open(args.keywords_file, "r", encoding="utf-8") as f:
                file_keywords = [line.strip() for line in f if line.strip()]
                args.keywords.extend(file_keywords)
        except Exception as e:
            print(f"⚠️ Failed to load keywords from {args.keywords_file}: {e}")

    link_results = []
    keyword_results = []
    bates_results = []
    file_results = []

    # Build a combined set of extensions to search for, plus a single regex.
    ext_set = set()
    if args.source_code:
        ext_set.update(SOURCE_CODE_EXTENSIONS)
    for e in args.file_ext:
        e = e.lstrip(".")
        if e:
            ext_set.add(e.lower())
    file_ext_regex = build_file_pattern(sorted(ext_set)) if ext_set else None

    if args.folder:
        pdf_files = []
        for root, _, files in os.walk(args.path):
            for f in files:
                if f.lower().endswith(".pdf"):
                    pdf_files.append(os.path.join(root, f))
    else:
        if not os.path.isfile(args.path):
            print(f"❌ {args.path} is not a file. Use --folder to scan a directory of PDFs instead.")
            return
        pdf_files = [args.path]

    for filepath in tqdm(pdf_files, desc="Processing PDFs", unit="file"):
        filename = os.path.basename(filepath)
        try:
            doc = fitz.open(filepath)

            for page_num, page in enumerate(doc, start=1):
                need_text = args.text_urls or args.keywords or args.bates_body or file_ext_regex is not None
                if need_text:
                    text = page.get_text("text")
                    # Clean full page text before matching (safe replacement for display only)
                    cleaned_text = text.replace('\r', '').replace('\n', ' ')

                # Extract per-page Bates number from the bottom-right footer
                bates_id_footer = None
                if args.bates_footer_prefix:
                    br_text = page.get_textbox((page.rect.width - 200, page.rect.height - 100, page.rect.width, page.rect.height))
                    footer_pattern = re.compile(rf"{re.escape(args.bates_footer_prefix)}[_\-]?\d+", re.IGNORECASE)
                    footer_matches = footer_pattern.findall(br_text)
                    if footer_matches:
                        bates_id_footer = footer_matches[-1]  # Last match in the footer

                bates_fields = {}
                if args.bates_footer_prefix:
                    bates_fields["Bates ID (Footer)"] = bates_id_footer

                # Every Bates-style reference number found anywhere in the page's
                # visible content, regardless of prefix, with surrounding context.
                # Citations may carry a dotted page suffix (a page or page range
                # offset into the root document, e.g. ACME0000456.542-550); that
                # suffix is captured separately so it can be merged into a
                # superset of cited page ranges in the report.
                if args.bates_body:
                    for match in generic_bates_regex.finditer(text):
                        bates_id = match.group(1)  # base Bates ID (no suffix)
                        match_start = match.start()
                        match_end = match.end()
                        context = clean_context_string(
                            extract_context(text, match_start, match_end, args.context_window)
                        )
                        suffix = (match.group(2) or "").lstrip(".")  # e.g. "" "8" "25-26" "542-550"
                        bates_results.append({
                            "Filename": filename,
                            "Page": page_num,
                            "Match Type": "bates_body",
                            "Bates ID": bates_id,
                            "Cited Page": suffix,
                            "Matched String": match.group(0),
                            "Context": context,
                            **bates_fields
                        })

                # File-name citations (source code and/or specific extensions)
                if file_ext_regex is not None:
                    for match in file_ext_regex.finditer(text):
                        match_start = match.start()
                        match_end = match.end()
                        context = clean_context_string(
                            extract_context(text, match_start, match_end, args.context_window)
                        )
                        file_results.append({
                            "Filename": filename,
                            "Page": page_num,
                            "Match Type": "file_citation",
                            "Matched String": match.group(),
                            "Context": context,
                            **bates_fields
                        })

                # Link annotations
                # --- Link annotations ---
                if args.link_annotations:
                    for link in page.get_links():
                        uri = link.get("uri")
                        if uri:
                            link_results.append({
                                "Filename": filename,
                                "Page": page_num,
                                "Match Type": "link",
                                "Matched String": uri,
                                "Context": "",
                                **bates_fields,
                                "Base URL": get_base_url(uri)
                            })


                # Text-based search
                if args.text_urls or args.keywords:
                    if args.text_urls:
                        for match in url_regex.finditer(cleaned_text):
                            match_start = match.start()
                            match_end = match.end()
                            context = clean_context_string(
                                extract_context(text, match_start, match_end, args.context_window)
                            )
                            raw_url = match.group()
                            url = clean_url_string(raw_url)
                            link_results.append({
                                "Filename": filename,
                                "Page": page_num,
                                "Match Type": "text_url",
                                "Matched String": url,
                                "Context": context,
                                **bates_fields,
                                "Base URL": get_base_url(url)
                            })

                    for keyword in args.keywords:
                        for match in re.finditer(re.escape(keyword), text, flags=re.IGNORECASE):
                            match_start = match.start()
                            match_end = match.end()
                            context = clean_context_string(
                                extract_context(text, match_start, match_end, args.context_window)
                            )
                            keyword_results.append({
                                "Filename": filename,
                                "Page": page_num,
                                "Match Type": "keyword",
                                "Matched String": match.group(),
                                "Context": context,
                                **bates_fields
                            })

            doc.close()

        except Exception as e:
            print(f"❌ Error processing {filename}: {e}")


    df_links=pd.DataFrame()
    df_keywords=pd.DataFrame()
    df_bates=pd.DataFrame()
    df_files=pd.DataFrame()

    if link_results:
        # Create dataframe from raw link results
        df_links_raw = pd.DataFrame(link_results)

        # Group by exact URL
        grouped_links = df_links_raw.groupby("Matched String")
        bates_columns = [c for c in ("Bates ID (Footer)",) if c in df_links_raw.columns]

        merged_link_rows = []

        for url, group in grouped_links:
            base_url = group["Base URL"].iloc[0]
            match_type = group["Match Type"].iloc[0]

            references = group.apply(
                lambda row: {
                    "Filename": row["Filename"],
                    "Page": row["Page"],
                    "Context": row["Context"],
                    **{col: row[col] for col in bates_columns}
                }, axis=1
            ).tolist()

            merged_link_rows.append({
                "Matched String": url,
                "Base URL": base_url,
                "Match Type": match_type,
                "Reference Count": len(references),
                "References": references
            })

        df_links = pd.DataFrame(merged_link_rows)
        df_links.sort_values(by=["Base URL", "Reference Count"], ascending=[True, False], inplace=True)

    if keyword_results:
        # Create dataframe from raw keyword results
        df_keywords_raw = pd.DataFrame(keyword_results)

        # Group by exact matched keyword
        grouped_keywords = df_keywords_raw.groupby("Matched String")
        bates_columns = [c for c in ("Bates ID (Footer)",) if c in df_keywords_raw.columns]

        merged_keyword_rows = []

        for keyword, group in grouped_keywords:
            match_type = group["Match Type"].iloc[0]

            references = group.apply(
                lambda row: {
                    "Filename": row["Filename"],
                    "Page": row["Page"],
                    "Context": row["Context"],
                    **{col: row[col] for col in bates_columns}
                }, axis=1
            ).tolist()

            merged_keyword_rows.append({
                "Matched String": keyword,
                "Match Type": match_type,
                "Reference Count": len(references),
                "References": references
            })

        df_keywords = pd.DataFrame(merged_keyword_rows)
        df_keywords.sort_values(by=["Reference Count", "Matched String"], ascending=[False, True], inplace=True)

    if bates_results:
        # Create dataframe from raw body Bates number results
        df_bates_raw = pd.DataFrame(bates_results)

        # Group by the base Bates ID, so that citations that differ only by
        # their page suffix (e.g. ACME0000123.8 and ACME0000123.25-26)
        # collapse onto a single row.
        grouped_bates = df_bates_raw.groupby("Bates ID")
        bates_columns = [c for c in ("Bates ID (Footer)",) if c in df_bates_raw.columns]

        merged_bates_rows = []

        for bates_id, group in grouped_bates:
            references = group.apply(
                lambda row: {
                    "Filename": row["Filename"],
                    "Page": row["Page"],
                    "Cited Page": row["Cited Page"],
                    "Matched String": row["Matched String"],
                    "Context": row["Context"],
                    **{col: row[col] for col in bates_columns}
                }, axis=1
            ).tolist()

            # Collect the page offset cited by each occurrence (only those that
            # carried a suffix) and merge them into superset ranges of pages
            # being cited.
            ranges = [
                r
                for r in (parse_page_range(ref["Cited Page"]) for ref in references)
                if r is not None
            ]
            merged_ranges = merge_page_ranges(ranges)

            merged_bates_rows.append({
                "Bates ID": bates_id,
                "Cited Page Range": format_merged_ranges(merged_ranges),
                "Reference Count": len(references),
                "References": references
            })

        df_bates = pd.DataFrame(merged_bates_rows)
        df_bates.sort_values(by=["Reference Count", "Bates ID"], ascending=[False, True], inplace=True)

    if file_results:
        # Create dataframe from raw file-citation results
        df_files_raw = pd.DataFrame(file_results)

        grouped_files = df_files_raw.groupby("Matched String")
        bates_columns = [c for c in ("Bates ID (Footer)",) if c in df_files_raw.columns]

        merged_file_rows = []

        for file_name, group in grouped_files:
            references = group.apply(
                lambda row: {
                    "Filename": row["Filename"],
                    "Page": row["Page"],
                    "Context": row["Context"],
                    **{col: row[col] for col in bates_columns}
                }, axis=1
            ).tolist()

            merged_file_rows.append({
                "Matched String": file_name,
                "Match Type": group["Match Type"].iloc[0],
                "Reference Count": len(references),
                "References": references
            })

        df_files = pd.DataFrame(merged_file_rows)
        df_files.sort_values(by=["Reference Count", "Matched String"], ascending=[False, True], inplace=True)

    json_folder = os.path.join(output_dir, "references_json")
    if not df_links.empty or not df_keywords.empty or not df_bates.empty or not df_files.empty:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            if not df_links.empty:
                df_links["References"] = df_links.apply(
                    lambda row: save_large_json(
                        row["References"],
                        base_filename=f"{row['Matched String'][:50].strip().replace('/', '_')}_links",
                        folder=json_folder
                    ),
                    axis=1
                )
                df_links.to_excel(writer, sheet_name="URLs", index=False)
            if not df_keywords.empty:
                df_keywords["References"] = df_keywords.apply(
                    lambda row: save_large_json(
                        row["References"],
                        base_filename=f"{row['Matched String'][:50].strip().replace('/', '_')}_keywords",
                        folder=json_folder
                    ),
                    axis=1
)
                df_keywords.to_excel(writer, sheet_name="Keywords", index=False)
            if not df_bates.empty:
                df_bates["References"] = df_bates.apply(
                    lambda row: save_large_json(
                        row["References"],
                        base_filename=f"{row['Bates ID'][:50].strip().replace('/', '_')}_bates",
                        folder=json_folder
                    ),
                    axis=1
                )
                df_bates.to_excel(writer, sheet_name="Bates Numbers", index=False)
            if not df_files.empty:
                df_files["References"] = df_files.apply(
                    lambda row: save_large_json(
                        row["References"],
                        base_filename=f"{row['Matched String'][:50].strip().replace('/', '_').replace(chr(92), '_')}_files",
                        folder=json_folder
                    ),
                    axis=1
                )
                df_files.to_excel(writer, sheet_name="File Citations", index=False)

        format_excel(excel_path)
        print(f"[OK] Excel saved to {excel_path}")
    else:
        print("[ERR] no references or links found")



if __name__ == "__main__":
    main()
